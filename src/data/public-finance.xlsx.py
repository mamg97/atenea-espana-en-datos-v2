import io
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _google_sheets import read_sheet_or_seed

BASE = Path(__file__).resolve().parent

FACT_RANGE = "Fact_finanzas_publicas!A:AC"
FACT_SEED = BASE / "seed" / "fact_finanzas_publicas_2024.csv"
SOURCE_RANGE = "Dim_fuentes!A:H"
SOURCE_SEED = BASE / "seed" / "dim_fuentes_finanzas_publicas.csv"
FIG_RANGE = "Figuras_ingresos_no_sumables!A:K"
FIG_SEED = BASE / "seed" / "figuras_ingresos_2024.csv"

fact_rows, fact_origin = read_sheet_or_seed(FACT_RANGE, FACT_SEED)
source_rows, source_origin = read_sheet_or_seed(SOURCE_RANGE, SOURCE_SEED)
fig_rows, fig_origin = read_sheet_or_seed(FIG_RANGE, FIG_SEED)

wb = Workbook()
ws = wb.active
ws.title = "README"

readme = [
    ["ATENEA — Finanzas públicas"],
    ["Libro generado automáticamente durante el build de la web."],
    [""],
    ["Tabla", "Origen usado en este build"],
    ["Fact_finanzas_publicas", fact_origin],
    ["Dim_fuentes", source_origin],
    ["Figuras_ingresos_no_sumables", fig_origin],
    [""],
    ["Nota metodológica"],
    [
        "Las figuras tributarias de la hoja Figuras_ingresos_no_sumables son informativas y no forman una lista exhaustiva; no deben sumarse para reconstruir el total S.13."
    ],
]
for row in readme:
    ws.append(row)

navy = "0B2447"
gold = "B89B5E"
cream = "F8F5EF"

ws["A1"].font = Font(bold=True, size=18, color=navy)
ws["A9"].font = Font(bold=True, color=navy)
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 24
ws.sheet_view.showGridLines = False


def add_table_sheet(name, rows):
    sheet = wb.create_sheet(name)
    for row in rows:
        sheet.append(list(row))

    if rows:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for col_idx, col_cells in enumerate(sheet.columns, start=1):
        values = [str(c.value or "") for c in list(col_cells)[:250]]
        max_len = max([len(v) for v in values] + [8])
        width = min(max(max_len + 2, 10), 55)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # Imported/raw data are linked inputs: show them in green text.
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="008000")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return sheet


add_table_sheet("fact_finanzas_publicas", fact_rows)
add_table_sheet("dim_fuentes", source_rows)
add_table_sheet("figuras_ingresos_no_sumables", fig_rows)

buffer = io.BytesIO()
wb.save(buffer)
sys.stdout.buffer.write(buffer.getvalue())
